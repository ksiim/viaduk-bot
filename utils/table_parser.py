import openpyxl
from abc import ABC, abstractmethod
from bot import bot
from models.dbs.orm import Orm


class BaseParser(ABC):
    """
    Base class for table parsers.
    """

    def __init__(self, file_path: str):
        """Initializes the parser with the file path.

        Args:
            file_path (str): The path to the file.
        """
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active

    @abstractmethod
    async def parse(self):
        """Parses the table and returns the data."""
        pass

    async def close(self):
        """Closes the workbook."""
        self.workbook.close()


class Parser(BaseParser):
    """
    Parser for full table.
    """

    async def parse(self):
        """Parses the full table and returns the data."""
        data = {row[0]: (row[1], row[2]) for row in self.sheet.iter_rows(values_only=True)}
        return data
    
    async def get_debt(self, garage_number: str):
        """Returns the debt for the garage number."""
        row = await self.get_row(garage_number)
        return row[0]
    
    async def get_payment(self, garage_number: str):
        """Returns the payment for the garage number"""
        row = await self.get_row(garage_number)
        return row[1]

    async def get_row(self, garage_number):
        """Returns the row by the garage number"""
        data = await self.parse()
        row = data.get(garage_number)
        return row
    
    async def update_debts(self):
        """Updates the debts table."""
        for index, row in enumerate(self.sheet.iter_rows(values_only=True)):
            if row[0] == None:
                break
            debt, additional_value = row[1], row[2]
            if debt == None or additional_value == None:
                continue
            try:
                new_debt = int(debt) - int(additional_value)
            except:
                admins = await Orm.get_admins()
                for admin in admins:
                    await bot.send_message(admin.telegram_id, f"Ошибка при обновлении в строке {index + 1} с долгом {debt} и дополнительным значением {additional_value}")
                new_debt = debt
            self.sheet.cell(row=index + 1, column=2, value=new_debt)
        self.workbook.save(self.file_path)
    